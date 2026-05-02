from __future__ import annotations

from datetime import date, datetime, time, timedelta
from io import BytesIO
import re

import altair as alt
import pandas as pd
import streamlit as st
import yfinance as yf


st.set_page_config(
    page_title="Manual Weekly Stock Signal Tracker",
    page_icon="📈",
    layout="wide",
)


st.markdown(
    """
    <style>
    [data-testid="stAppViewContainer"], .stApp, html, body {
        background: #020817 !important;
        color: #e5e7eb !important;
    }

    [data-testid="stHeader"] {
        background: rgba(2, 8, 23, 0.95) !important;
    }

    .block-container {
        padding-top: 2rem;
        padding-bottom: 2rem;
        max-width: 1300px;
    }

    h1, h2, h3, h4, h5, h6 {
        color: #f8fafc !important;
    }

    p, label, small, .stCaption, [data-testid="stCaptionContainer"] {
        color: #cbd5e1 !important;
    }

    [data-testid="stMarkdownContainer"] p,
    [data-testid="stMarkdownContainer"] li,
    [data-testid="stMarkdownContainer"] span {
        color: #cbd5e1 !important;
    }

    .chart-wrap {
        background: #0f172a;
        border: 1px solid #334155;
        border-radius: 24px;
        padding: 20px 20px 12px 20px;
        box-shadow: 0 10px 28px rgba(0, 0, 0, 0.30);
        margin-bottom: 24px;
    }

    .chart-title-row {
        display: flex;
        justify-content: space-between;
        align-items: flex-start;
        gap: 18px;
        margin-bottom: 4px;
    }

    .chart-big-number {
        font-size: 36px;
        font-weight: 850;
        letter-spacing: -1.1px;
        color: #f8fafc;
        line-height: 1;
        margin-bottom: 6px;
    }

    .chart-change-green {
        color: #34d399;
        font-size: 15px;
        font-weight: 850;
    }

    .chart-change-red {
        color: #f87171;
        font-size: 15px;
        font-weight: 850;
    }

    .chart-small-note {
        color: #cbd5e1;
        font-size: 13px;
        font-weight: 750;
        text-align: right;
    }

    div[role="radiogroup"] {
        background: #111827;
        border: 1px solid #334155;
        padding: 4px;
        border-radius: 999px;
        width: fit-content;
        margin-bottom: 10px;
    }

    div[role="radiogroup"] label {
        padding: 4px 10px;
        border-radius: 999px;
        color: #cbd5e1 !important;
        font-weight: 800;
    }

    div[role="radiogroup"] label:has(input:checked) {
        background: #1f2937;
        color: #f8fafc !important;
        box-shadow: 0 1px 5px rgba(0, 0, 0, 0.25);
    }

    [data-testid="stMetricLabel"] {
        color: #cbd5e1 !important;
        font-weight: 750 !important;
    }

    [data-testid="stMetricValue"] {
        color: #f8fafc !important;
        font-weight: 850 !important;
    }

    [data-testid="stMetricDelta"] {
        color: #e5e7eb !important;
        font-weight: 800 !important;
    }

    [data-baseweb="tab-list"] button {
        color: #cbd5e1 !important;
    }

    [data-baseweb="tab-list"] button[aria-selected="true"] {
        color: #f8fafc !important;
    }

    [data-baseweb="select"] > div,
    [data-testid="stNumberInput"] input,
    [data-testid="stTextArea"] textarea,
    [data-testid="stTextInput"] input {
        background: #111827 !important;
        color: #f8fafc !important;
        border-color: #475569 !important;
    }

    [data-testid="stSelectbox"] label,
    [data-testid="stNumberInput"] label,
    [data-testid="stTextArea"] label,
    [data-testid="stTextInput"] label {
        color: #e5e7eb !important;
        font-weight: 700 !important;
    }

    .stDownloadButton button,
    .stButton button {
        background: #111827 !important;
        color: #f8fafc !important;
        border: 1px solid #475569 !important;
    }

    .stDownloadButton button:hover,
    .stButton button:hover {
        background: #1f2937 !important;
        color: #ffffff !important;
        border-color: #64748b !important;
    }

    div[data-testid="stDataFrame"] {
        border: 1px solid #334155;
        border-radius: 14px;
        overflow: hidden;
    }
    </style>
    """,
    unsafe_allow_html=True,
)


DEFAULT_TEXT = """1   PHM       SELL      DOWN      95          STRONG    NORMAL      -1.000    1.43      -0.000    2026-04-27T22:58:34+00:00
2   EXPE      SELL      DOWN      88          STRONG    EVENTFUL    -0.970    -4.18     -0.000    2026-04-27T22:43:27+00:00
3   FCX       SELL      DOWN      95          STRONG    EVENTFUL    -1.000    0.41      -0.000    2026-04-27T22:44:58+00:00
4   TSLA      SELL      DOWN      67          MODERATE  NORMAL      -1.000    0.72      -0.000    2026-04-27T23:02:44+00:00
5   XOM       SELL      DOWN      61          MODERATE  NORMAL      -0.638    1.63      -0.000    2026-04-27T22:43:41+00:00
6   VLTO      SELL      DOWN      54          MODERATE  NORMAL      -0.560    -2.82     -0.000    2026-04-27T23:04:54+00:00
7   BLK       BUY       UP        51          MODERATE  NORMAL      0.535     4.22      -0.000    2026-04-27T22:34:05+00:00
8   MCO       BUY       UP        48          MODERATE  NORMAL      0.495     3.46      -0.000    2026-04-27T22:53:56+00:00
9   SMCI      SELL      DOWN      50          MODERATE  NORMAL      -0.631    -7.38     -0.000    2026-04-27T23:01:47+00:00
10  RCL       SELL      DOWN      45          MODERATE  NORMAL      -0.472    -3.04     -0.000    2026-04-27T23:00:02+00:00
11  BA        BUY       UP        46          MODERATE  NORMAL      0.444     3.27      -0.000    2026-04-27T22:34:22+00:00
12  NCLH      SELL      DOWN      50          MODERATE  EVENTFUL    -0.514    -2.80     -0.000    2026-04-27T22:55:20+00:00
13  VST       WATCH     UP        42          WEAK      NORMAL      0.455     4.64      -0.000    2026-04-27T23:05:39+00:00
14  COF       WATCH     UP        40          WEAK      NORMAL      0.428     5.55      -0.000    2026-04-27T22:35:36+00:00
15  CARR      WATCH     UP        43          WEAK      NORMAL      0.452     3.17      -0.000    2026-04-27T22:35:49+00:00
16  MSCI      WATCH     UP        43          WEAK      NORMAL      0.443     2.87      -0.000    2026-04-27T22:54:18+00:00
17  APTV      WATCH     UP        39          WEAK      NORMAL      0.407     4.83      -0.000    2026-04-27T22:32:17+00:00
18  EL        WATCH     UP        39          WEAK      NORMAL      0.389     4.99      -0.000    2026-04-27T22:43:01+00:00
19  ALLE      WATCH     UP        35          WEAK      NORMAL      0.421     6.07      -0.000    2026-04-27T22:30:17+00:00
20  DOV       WATCH     UP        43          WEAK      NORMAL      0.439     2.22      -0.000    2026-04-27T22:41:12+00:00
21  ARE       WATCH     UP        38          WEAK      NORMAL      0.385     4.88      -0.000    2026-04-27T22:30:08+00:00
22  AXP       WATCH     UP        38          WEAK      NORMAL      0.396     4.55      -0.000    2026-04-27T22:31:06+00:00
23  ROK       WATCH     UP        39          WEAK      NORMAL      0.418     3.78      -0.000    2026-04-27T22:59:44+00:00
24  RF        WATCH     UP        42          WEAK      NORMAL      0.394     2.80      -0.000    2026-04-27T22:59:22+00:00
25  NFLX      WATCH     DOWN      35          WEAK      NORMAL      -0.377    -5.86     -0.000    2026-04-27T22:54:32+00:00"""



def this_weeks_monday(today: date | None = None) -> date:
    today = today or date.today()
    return today - timedelta(days=today.weekday())


def yahoo_symbol(ticker: str) -> str:
    return str(ticker).strip().upper().replace(".", "-")


def parse_model_output(raw_text: str) -> pd.DataFrame:
    rows = []

    for line in raw_text.splitlines():
        line = line.strip()
        if not line or line.lower().startswith("rank"):
            continue

        parts = re.split(r"\s+", line)

        if len(parts) < 11:
            continue

        try:
            rows.append(
                {
                    "Rank": int(parts[0]),
                    "Ticker": parts[1].upper(),
                    "Action": parts[2].upper(),
                    "Direction": parts[3].upper(),
                    "Conviction": int(float(parts[4])),
                    "Edge": parts[5].upper(),
                    "Regime": parts[6].upper(),
                    "Model Score": float(parts[7]),
                    "Expected Move %": float(parts[8]),
                    "Setup Score": float(parts[9]),
                    "Run Timestamp": parts[10],
                }
            )
        except Exception:
            continue

    return pd.DataFrame(rows)


@st.cache_data(ttl=300, show_spinner=False)
def fetch_prices(tickers: tuple[str, ...], monday_date: date) -> pd.DataFrame:
    records = []
    start_date = monday_date
    end_date = datetime.now().date() + timedelta(days=1)

    for ticker in tickers:
        symbol = yahoo_symbol(ticker)
        monday_price = None
        current_price = None
        source = "N/A"
        error = ""

        try:
            intraday = yf.download(
                symbol,
                start=start_date.isoformat(),
                end=end_date.isoformat(),
                interval="30m",
                auto_adjust=True,
                progress=False,
                prepost=False,
                threads=False,
            )

            if intraday is not None and not intraday.empty:
                if isinstance(intraday.columns, pd.MultiIndex):
                    intraday.columns = [c[0] for c in intraday.columns]

                intraday = intraday.rename(columns=str.lower)
                intraday.index = pd.to_datetime(intraday.index).tz_localize(None)

                monday_rows = intraday[intraday.index.date == monday_date]

                if not monday_rows.empty and "close" in monday_rows.columns:
                    window_start = datetime.combine(monday_date, time(11, 30))
                    window_end = datetime.combine(monday_date, time(12, 30))
                    noon_window = monday_rows[
                        (monday_rows.index >= window_start)
                        & (monday_rows.index <= window_end)
                    ]

                    if not noon_window.empty:
                        monday_price = float(noon_window["close"].dropna().mean())
                        source = "Monday 11:30-12:30 avg"
                    elif not monday_rows["close"].dropna().empty:
                        monday_price = float(monday_rows["close"].dropna().iloc[0])
                        source = "Monday first intraday price"

                if "close" in intraday.columns and not intraday["close"].dropna().empty:
                    current_price = float(intraday["close"].dropna().iloc[-1])

            if monday_price is None or current_price is None:
                daily = yf.download(
                    symbol,
                    start=start_date.isoformat(),
                    end=end_date.isoformat(),
                    interval="1d",
                    auto_adjust=True,
                    progress=False,
                    threads=False,
                )

                if daily is not None and not daily.empty:
                    if isinstance(daily.columns, pd.MultiIndex):
                        daily.columns = [c[0] for c in daily.columns]

                    daily = daily.rename(columns=str.lower)

                    if "close" in daily.columns and not daily["close"].dropna().empty:
                        if monday_price is None:
                            monday_price = float(daily["close"].dropna().iloc[0])
                            source = "Monday daily close fallback"

                        if current_price is None:
                            current_price = float(daily["close"].dropna().iloc[-1])

        except Exception as exc:
            error = str(exc)

        records.append(
            {
                "Ticker": ticker,
                "Monday Reference Price": monday_price,
                "Current Price": current_price,
                "Reference Price Source": source,
                "Price Error": error,
            }
        )

    return pd.DataFrame(records)


@st.cache_data(ttl=300, show_spinner=False)
def fetch_position_return_series(model_rows: tuple[tuple[str, str, str], ...], monday_date: date) -> pd.DataFrame:
    """
    Builds Robinhood-style portfolio lines.

    BUY/UP = long return
    SELL/DOWN = short return, so returns are inverted
    WATCH is ignored
    """
    if not model_rows:
        return pd.DataFrame()

    start_date = monday_date
    end_date = datetime.now().date() + timedelta(days=1)

    buy_series = []
    sell_series = []

    for ticker, action, direction in model_rows:
        action = str(action).upper()
        direction = str(direction).upper()

        if not ((action == "BUY" and direction == "UP") or (action == "SELL" and direction == "DOWN")):
            continue

        symbol = yahoo_symbol(ticker)

        try:
            df = yf.download(
                symbol,
                start=start_date.isoformat(),
                end=end_date.isoformat(),
                interval="30m",
                auto_adjust=True,
                progress=False,
                prepost=False,
                threads=False,
            )

            if df is None or df.empty:
                continue

            if isinstance(df.columns, pd.MultiIndex):
                df.columns = [c[0] for c in df.columns]

            df = df.rename(columns=str.lower)
            df.index = pd.to_datetime(df.index).tz_localize(None)

            if "close" not in df.columns or df["close"].dropna().empty:
                continue

            monday_rows = df[df.index.date == monday_date]
            if monday_rows.empty:
                continue

            window_start = datetime.combine(monday_date, time(11, 30))
            window_end = datetime.combine(monday_date, time(12, 30))
            noon_window = monday_rows[
                (monday_rows.index >= window_start)
                & (monday_rows.index <= window_end)
            ]

            if not noon_window.empty:
                ref_price = float(noon_window["close"].dropna().mean())
            else:
                ref_price = float(monday_rows["close"].dropna().iloc[0])

            long_pct = (df["close"] - ref_price) / ref_price * 100

            if action == "BUY" and direction == "UP":
                long_pct.name = ticker
                buy_series.append(long_pct)

            if action == "SELL" and direction == "DOWN":
                short_pct = -long_pct
                short_pct.name = ticker
                sell_series.append(short_pct)

        except Exception:
            continue

    chart_parts = []

    if buy_series:
        buy_df = pd.concat(buy_series, axis=1).sort_index()
        chart_parts.append(buy_df.mean(axis=1).rename("BUY basket"))

    if sell_series:
        sell_df = pd.concat(sell_series, axis=1).sort_index()
        chart_parts.append(sell_df.mean(axis=1).rename("SELL short basket"))

    if buy_series or sell_series:
        all_series = buy_series + sell_series
        combined_df = pd.concat(all_series, axis=1).sort_index()
        chart_parts.append(combined_df.mean(axis=1).rename("Combined active basket"))

    if not chart_parts:
        return pd.DataFrame()

    chart_df = pd.concat(chart_parts, axis=1).sort_index()
    return chart_df.dropna(how="all")


@st.cache_data(ttl=300, show_spinner=False)
def fetch_week_price_paths(tickers: tuple[str, ...], monday_date: date) -> pd.DataFrame:
    """
    Gets intraday close prices for every ticker from this week's Monday to now.
    Used to check whether each prediction became true at any point during the week.
    """
    records = []
    start_date = monday_date
    end_date = datetime.now().date() + timedelta(days=1)

    for ticker in tickers:
        symbol = yahoo_symbol(ticker)

        try:
            df = yf.download(
                symbol,
                start=start_date.isoformat(),
                end=end_date.isoformat(),
                interval="30m",
                auto_adjust=True,
                progress=False,
                prepost=False,
                threads=False,
            )

            if df is None or df.empty:
                continue

            if isinstance(df.columns, pd.MultiIndex):
                df.columns = [c[0] for c in df.columns]

            df = df.rename(columns=str.lower)
            df.index = pd.to_datetime(df.index).tz_localize(None)

            if "close" not in df.columns:
                continue

            path = df[["close"]].dropna().copy()
            path["Ticker"] = ticker
            path["Time"] = path.index
            path = path.rename(columns={"close": "Price"})

            records.append(path[["Ticker", "Time", "Price"]])

        except Exception:
            continue

    if not records:
        return pd.DataFrame(columns=["Ticker", "Time", "Price"])

    return pd.concat(records, ignore_index=True)


def build_weekly_price_tracker(
    tracker_df: pd.DataFrame,
    path_df: pd.DataFrame,
) -> pd.DataFrame:
    """
    Checks the whole week's intraday path for each stock.

    For UP:
      - correct if price traded above Monday reference price
      - best correct price is weekly high

    For DOWN:
      - correct if price traded below Monday reference price
      - best correct price is weekly low

    For NEUTRAL:
      - correct if it stayed within +/- 0.50% of Monday reference price
      - best correct price is the closest price to Monday reference price

    Money logic:
      - BUY/UP makes money at best correct price: best price - Monday price
      - SELL/DOWN makes money at best correct price: Monday price - best price
      - never-correct stocks get $0 for best-correct P/L
    """
    output_rows = []

    if tracker_df.empty:
        return pd.DataFrame()

    for _, row in tracker_df.iterrows():
        ticker = str(row.get("Ticker", "")).upper()
        predicted = str(row.get("Direction", "")).upper()
        action = str(row.get("Action", "")).upper()

        monday_price = row.get("Monday Reference Price")
        stock_path = path_df[path_df["Ticker"] == ticker].copy()

        base = {
            "Ticker": ticker,
            "Action": action,
            "Predicted Direction": predicted,
            "Prediction True During Week": "N/A",
            "Monday Price": monday_price,
            "First Correct Time": None,
            "Best Correct Time": None,
            "Best Correct Price": None,
            "Best Correct Move %": None,
            "1-Share Best Correct P/L": 0.0,
            "Final Price Used": None,
            "Final Move %": None,
            "Final 1-Share P/L": 0.0,
        }

        if pd.isna(monday_price) or stock_path.empty:
            output_rows.append(base)
            continue

        monday_price = float(monday_price)

        stock_path["Move %"] = (stock_path["Price"] - monday_price) / monday_price * 100
        stock_path = stock_path.sort_values("Time")

        final_row = stock_path.iloc[-1]
        final_price = float(final_row["Price"])
        final_move_pct = float(final_row["Move %"])

        base["Final Price Used"] = final_price
        base["Final Move %"] = final_move_pct

        max_idx = stock_path["Price"].idxmax()
        min_idx = stock_path["Price"].idxmin()
        highest = stock_path.loc[max_idx]
        lowest = stock_path.loc[min_idx]

        if predicted == "UP":
            true_rows = stock_path[stock_path["Price"] > monday_price]

            if not true_rows.empty:
                best = highest
                best_price = float(best["Price"])
                pnl = best_price - monday_price

                base["Prediction True During Week"] = "YES"
                base["First Correct Time"] = true_rows.iloc[0]["Time"]
                base["Best Correct Time"] = best["Time"]
                base["Best Correct Price"] = best_price
                base["Best Correct Move %"] = float(best["Move %"])
                base["1-Share Best Correct P/L"] = pnl
            else:
                base["Prediction True During Week"] = "NO"

            base["Final 1-Share P/L"] = final_price - monday_price

        elif predicted == "DOWN":
            true_rows = stock_path[stock_path["Price"] < monday_price]

            if not true_rows.empty:
                best = lowest
                best_price = float(best["Price"])
                pnl = monday_price - best_price

                base["Prediction True During Week"] = "YES"
                base["First Correct Time"] = true_rows.iloc[0]["Time"]
                base["Best Correct Time"] = best["Time"]
                base["Best Correct Price"] = best_price
                base["Best Correct Move %"] = float(best["Move %"])
                base["1-Share Best Correct P/L"] = pnl
            else:
                base["Prediction True During Week"] = "NO"

            base["Final 1-Share P/L"] = monday_price - final_price

        elif predicted == "NEUTRAL":
            true_rows = stock_path[stock_path["Move %"].abs() <= 0.50]

            if not true_rows.empty:
                closest_idx = true_rows["Move %"].abs().idxmin()
                best = true_rows.loc[closest_idx]

                base["Prediction True During Week"] = "YES"
                base["First Correct Time"] = true_rows.iloc[0]["Time"]
                base["Best Correct Time"] = best["Time"]
                base["Best Correct Price"] = float(best["Price"])
                base["Best Correct Move %"] = float(best["Move %"])
                base["1-Share Best Correct P/L"] = 0.0
            else:
                base["Prediction True During Week"] = "NO"

            base["Final 1-Share P/L"] = 0.0

        output_rows.append(base)

    out = pd.DataFrame(output_rows)

    if out.empty:
        return out

    preferred_cols = [
        "Ticker",
        "Action",
        "Predicted Direction",
        "Prediction True During Week",
        "Monday Price",
        "First Correct Time",
        "Best Correct Time",
        "Best Correct Price",
        "Best Correct Move %",
        "1-Share Best Correct P/L",
        "Final Price Used",
        "Final Move %",
        "Final 1-Share P/L",
    ]

    return out[preferred_cols]


def build_weekly_truth_summary(path_tracker_df: pd.DataFrame) -> dict:
    if path_tracker_df.empty:
        return {
            "total": 0,
            "true_count": 0,
            "false_count": 0,
            "true_pct": 0.0,
            "best_correct_pnl": 0.0,
            "wrong_final_pnl": 0.0,
            "perfect_exit_minus_wrong_pnl": 0.0,
            "final_pnl": 0.0,
        }

    valid = path_tracker_df[
        path_tracker_df["Prediction True During Week"].isin(["YES", "NO"])
    ].copy()

    if valid.empty:
        return {
            "total": 0,
            "true_count": 0,
            "false_count": 0,
            "true_pct": 0.0,
            "best_correct_pnl": 0.0,
            "wrong_final_pnl": 0.0,
            "perfect_exit_minus_wrong_pnl": 0.0,
            "final_pnl": 0.0,
        }

    true_mask = valid["Prediction True During Week"] == "YES"
    false_mask = valid["Prediction True During Week"] == "NO"

    true_count = int(true_mask.sum())
    false_count = int(false_mask.sum())
    total = int(len(valid))
    true_pct = true_count / total * 100 if total else 0.0

    # Correct stocks: assume you sold/covered 1 share at the best correct price.
    best_correct_pnl = float(valid.loc[true_mask, "1-Share Best Correct P/L"].fillna(0).sum())

    # Wrong stocks: they were never correct, so count their latest/final result.
    wrong_final_pnl = float(valid.loc[false_mask, "Final 1-Share P/L"].fillna(0).sum())

    # This is the dashboard number you asked for:
    # perfect exit when right, minus/add the final result when wrong.
    perfect_exit_minus_wrong_pnl = best_correct_pnl + wrong_final_pnl

    # This is just the simple held-to-now result for comparison.
    final_pnl = float(valid["Final 1-Share P/L"].fillna(0).sum())

    return {
        "total": total,
        "true_count": true_count,
        "false_count": false_count,
        "true_pct": true_pct,
        "best_correct_pnl": best_correct_pnl,
        "wrong_final_pnl": wrong_final_pnl,
        "perfect_exit_minus_wrong_pnl": perfect_exit_minus_wrong_pnl,
        "final_pnl": final_pnl,
    }

def build_weekly_truth_group_summary(path_tracker_df: pd.DataFrame) -> pd.DataFrame:
    if path_tracker_df.empty:
        return pd.DataFrame(
            columns=[
                "Group",
                "Stocks",
                "1-Share Best Correct P/L",
                "Wrong/Final 1-Share P/L",
                "Best Exit - Wrong P/L",
                "Final 1-Share P/L",
            ]
        )

    valid = path_tracker_df[
        path_tracker_df["Prediction True During Week"].isin(["YES", "NO"])
    ].copy()

    if valid.empty:
        return pd.DataFrame(
            columns=[
                "Group",
                "Stocks",
                "1-Share Best Correct P/L",
                "Wrong/Final 1-Share P/L",
                "Best Exit - Wrong P/L",
                "Final 1-Share P/L",
            ]
        )

    rows = []

    for group_name, mask in [
        ("True during week", valid["Prediction True During Week"] == "YES"),
        ("False during week", valid["Prediction True During Week"] == "NO"),
        ("All stocks", valid["Prediction True During Week"].isin(["YES", "NO"])),
    ]:
        group = valid[mask].copy()

        true_group = group[group["Prediction True During Week"] == "YES"]
        false_group = group[group["Prediction True During Week"] == "NO"]

        best_correct = float(true_group["1-Share Best Correct P/L"].fillna(0).sum())
        wrong_final = float(false_group["Final 1-Share P/L"].fillna(0).sum())
        final_total = float(group["Final 1-Share P/L"].fillna(0).sum())

        rows.append(
            {
                "Group": group_name,
                "Stocks": int(len(group)),
                "1-Share Best Correct P/L": best_correct,
                "Wrong/Final 1-Share P/L": wrong_final,
                "Best Exit - Wrong P/L": best_correct + wrong_final,
                "Final 1-Share P/L": final_total,
            }
        )

    return pd.DataFrame(rows)

def style_weekly_path_tracker(df: pd.DataFrame):
    format_map = {
        "Monday Price": "${:,.2f}",
        "Best Correct Price": "${:,.2f}",
        "Best Correct Move %": "{:+.2f}%",
        "1-Share Best Correct P/L": "${:+,.2f}",
        "Final Price Used": "${:,.2f}",
        "Final Move %": "{:+.2f}%",
        "Final 1-Share P/L": "${:+,.2f}",
    }

    for col in ["First Correct Time", "Best Correct Time"]:
        if col in df.columns:
            format_map[col] = lambda x: "" if pd.isna(x) else pd.to_datetime(x).strftime("%a %I:%M %p")

    def color_yes_no(value):
        value = str(value).upper()

        if value == "YES":
            return "background-color: #047857; color: #ffffff; font-weight: 850;"
        if value == "NO":
            return "background-color: #b91c1c; color: #ffffff; font-weight: 850;"
        if value == "N/A":
            return "background-color: #374151; color: #ffffff; font-weight: 850;"

        return "color: #f8fafc; font-weight: 700;"

    def color_action(value):
        value = str(value).upper()

        if value == "BUY":
            return "background-color: #047857; color: #ffffff; font-weight: 850;"
        if value == "SELL":
            return "background-color: #b91c1c; color: #ffffff; font-weight: 850;"
        if value == "WATCH":
            return "background-color: #92400e; color: #ffffff; font-weight: 850;"

        return "color: #f8fafc; font-weight: 700;"

    def color_direction(value):
        value = str(value).upper()

        if value == "UP":
            return "background-color: #bbf7d0; color: #052e16; font-weight: 850;"
        if value == "DOWN":
            return "background-color: #fecaca; color: #450a0a; font-weight: 850;"
        if value == "NEUTRAL":
            return "background-color: #d1d5db; color: #111827; font-weight: 850;"

        return "color: #f8fafc; font-weight: 700;"

    def color_money(value):
        try:
            if value > 0:
                return "color: #34d399; font-weight: 850;"
            if value < 0:
                return "color: #f87171; font-weight: 850;"
        except Exception:
            pass

        return "color: #e5e7eb; font-weight: 700;"

    styled = df.style.format(format_map, na_rep="")

    if "Prediction True During Week" in df.columns:
        styled = styled.map(color_yes_no, subset=["Prediction True During Week"])

    if "Action" in df.columns:
        styled = styled.map(color_action, subset=["Action"])

    if "Predicted Direction" in df.columns:
        styled = styled.map(color_direction, subset=["Predicted Direction"])

    for col in [
        "Best Correct Move %",
        "1-Share Best Correct P/L",
        "Final Move %",
        "Final 1-Share P/L",
    ]:
        if col in df.columns:
            styled = styled.map(color_money, subset=[col])

    return styled

def filter_chart_range(chart_df: pd.DataFrame, selected_range: str) -> pd.DataFrame:
    if chart_df.empty:
        return chart_df

    chart_df = chart_df.copy()
    chart_df.index = pd.to_datetime(chart_df.index).tz_localize(None)
    latest_time = chart_df.index.max()

    if selected_range == "LIVE":
        start = latest_time - timedelta(hours=6)
        return chart_df[chart_df.index >= start]

    if selected_range == "1D":
        start = latest_time - timedelta(days=1)
        return chart_df[chart_df.index >= start]

    if selected_range == "2D":
        start = latest_time - timedelta(days=2)
        return chart_df[chart_df.index >= start]

    if selected_range == "1W":
        start = latest_time - timedelta(days=7)
        return chart_df[chart_df.index >= start]

    return chart_df


def add_tracking_columns(model_df: pd.DataFrame, price_df: pd.DataFrame) -> pd.DataFrame:
    out = model_df.merge(price_df, on="Ticker", how="left")

    out["Change Since Monday %"] = (
        (out["Current Price"] - out["Monday Reference Price"])
        / out["Monday Reference Price"]
        * 100
    )

    def actual_direction(change):
        if pd.isna(change):
            return "N/A"
        if change > 0:
            return "UP"
        if change < 0:
            return "DOWN"
        return "FLAT"

    out["Actual Direction So Far"] = out["Change Since Monday %"].apply(actual_direction)

    def correct(row):
        predicted = row["Direction"]
        actual = row["Actual Direction So Far"]

        if actual == "N/A":
            return "N/A"

        if predicted == "UP" and actual == "UP":
            return "YES"
        if predicted == "DOWN" and actual == "DOWN":
            return "YES"
        if predicted == "NEUTRAL" and actual == "FLAT":
            return "YES"

        return "NO"

    out["Correct So Far"] = out.apply(correct, axis=1)
    return out


def build_short_calculator_table(tracker_df: pd.DataFrame) -> pd.DataFrame:
    return tracker_df[
        (tracker_df["Action"] == "SELL")
        & (tracker_df["Direction"] == "DOWN")
        & (tracker_df["Monday Reference Price"].notna())
        & (tracker_df["Current Price"].notna())
    ].copy()


def build_what_if(tracker_df: pd.DataFrame) -> pd.DataFrame:
    """
    One-share what-if summary.

    BUY/UP = buy 1 share at Monday reference price.
    SELL/DOWN = short 1 share at Monday reference price.
    WATCH is ignored.
    """
    strategies = [
        (
            "BUY 1 share of each BUY/UP stock",
            (tracker_df["Action"] == "BUY") & (tracker_df["Direction"] == "UP"),
        ),
        (
            "SHORT 1 share of each SELL/DOWN stock",
            (tracker_df["Action"] == "SELL") & (tracker_df["Direction"] == "DOWN"),
        ),
        (
            "Combined active calls: 1 share each BUY/UP + SELL/DOWN",
            (
                ((tracker_df["Action"] == "BUY") & (tracker_df["Direction"] == "UP"))
                | ((tracker_df["Action"] == "SELL") & (tracker_df["Direction"] == "DOWN"))
            ),
        ),
    ]

    rows = []

    for name, mask in strategies:
        basket = tracker_df[mask].copy()
        total_entry_value = 0.0
        total_current_value = 0.0
        total_pnl = 0.0
        valid_positions = 0

        for _, row in basket.iterrows():
            monday_price = row.get("Monday Reference Price")
            current_price = row.get("Current Price")

            if pd.isna(monday_price) or pd.isna(current_price):
                continue

            monday_price = float(monday_price)
            current_price = float(current_price)

            action = str(row.get("Action", "")).upper()
            direction = str(row.get("Direction", "")).upper()

            if action == "BUY" and direction == "UP":
                pnl = current_price - monday_price
                current_position_value = current_price
            elif action == "SELL" and direction == "DOWN":
                pnl = monday_price - current_price
                current_position_value = monday_price + pnl
            else:
                continue

            total_entry_value += monday_price
            total_current_value += current_position_value
            total_pnl += pnl
            valid_positions += 1

        percent_return = (total_pnl / total_entry_value * 100) if total_entry_value else None

        rows.append(
            {
                "Strategy": name,
                "Stocks": valid_positions,
                "Monday Entry Value": total_entry_value if valid_positions else None,
                "Current Position Value": total_current_value if valid_positions else None,
                "Dollar P/L": total_pnl if valid_positions else None,
                "Percent Return": percent_return,
            }
        )

    return pd.DataFrame(rows)


def build_what_if_positions(tracker_df: pd.DataFrame, side: str) -> pd.DataFrame:
    """
    One-share detail table.

    BUY/UP = buy 1 share at Monday reference price.
    SELL/DOWN = short 1 share at Monday reference price.
    """
    side = side.upper()

    if side == "BUY":
        basket = tracker_df[
            (tracker_df["Action"] == "BUY")
            & (tracker_df["Direction"] == "UP")
            & (tracker_df["Change Since Monday %"].notna())
        ].copy()
        position_type = "Long"
    elif side == "SELL":
        basket = tracker_df[
            (tracker_df["Action"] == "SELL")
            & (tracker_df["Direction"] == "DOWN")
            & (tracker_df["Change Since Monday %"].notna())
        ].copy()
        position_type = "Short"
    else:
        basket = tracker_df[
            (
                ((tracker_df["Action"] == "BUY") & (tracker_df["Direction"] == "UP"))
                | ((tracker_df["Action"] == "SELL") & (tracker_df["Direction"] == "DOWN"))
            )
            & (tracker_df["Change Since Monday %"].notna())
        ].copy()
        position_type = "Mixed"

    if basket.empty:
        return pd.DataFrame(
            columns=[
                "Ticker",
                "Position",
                "Direction",
                "Monday Price",
                "Current Price",
                "Shares / Shorted Shares",
                "Stock Move %",
                "Position Return %",
                "Dollar P/L",
                "Correct So Far",
            ]
        )

    rows = []

    for _, row in basket.iterrows():
        action = str(row.get("Action", "")).upper()
        direction = str(row.get("Direction", "")).upper()
        stock_move = float(row["Change Since Monday %"])

        monday_price = row.get("Monday Reference Price")
        current_price = row.get("Current Price")

        if pd.isna(monday_price) or pd.isna(current_price):
            continue

        monday_price = float(monday_price)
        current_price = float(current_price)

        shares = 1.0

        if action == "BUY" and direction == "UP":
            pos_return = stock_move
            pos_label = "Long"
            pnl = current_price - monday_price
        elif action == "SELL" and direction == "DOWN":
            pos_return = -stock_move
            pos_label = "Short"
            pnl = monday_price - current_price
        else:
            continue

        rows.append(
            {
                "Ticker": row["Ticker"],
                "Position": pos_label if position_type == "Mixed" else position_type,
                "Direction": direction,
                "Monday Price": monday_price,
                "Current Price": current_price,
                "Shares / Shorted Shares": shares,
                "Stock Move %": stock_move,
                "Position Return %": pos_return,
                "Dollar P/L": pnl,
                "Correct So Far": row.get("Correct So Far"),
            }
        )

    return pd.DataFrame(rows)


def build_summary(df: pd.DataFrame, group_col: str) -> pd.DataFrame:
    # WATCH is intentionally excluded from summaries.
    active = df[df["Action"].isin(["BUY", "SELL"])].copy()
    valid = active[active["Correct So Far"].isin(["YES", "NO"])].copy()

    if valid.empty:
        return pd.DataFrame(columns=[group_col, "Stocks", "Correct", "Accuracy %", "Avg Change Since Monday %"])

    valid["Correct Flag"] = (valid["Correct So Far"] == "YES").astype(int)

    out = (
        valid.groupby(group_col)
        .agg(
            Stocks=("Ticker", "count"),
            Correct=("Correct Flag", "sum"),
            **{"Avg Change Since Monday %": ("Change Since Monday %", "mean")},
        )
        .reset_index()
    )

    out["Accuracy %"] = out["Correct"] / out["Stocks"] * 100
    return out[[group_col, "Stocks", "Correct", "Accuracy %", "Avg Change Since Monday %"]]


def make_portfolio_chart(filtered_chart_df: pd.DataFrame, chart_range: str):
    chart_data = filtered_chart_df.reset_index()

    first_col = chart_data.columns[0]
    chart_data = chart_data.rename(columns={first_col: "Time"})

    chart_data = (
        chart_data
        .melt(id_vars="Time", var_name="Basket", value_name="Return %")
        .dropna()
    )

    if chart_data.empty:
        return None, None, None

    latest_rows = (
        chart_data.sort_values("Time")
        .groupby("Basket", as_index=False)
        .tail(1)
        .sort_values("Basket")
    )

    combined_latest = latest_rows[latest_rows["Basket"] == "Combined active basket"]

    if not combined_latest.empty:
        latest_return = float(combined_latest["Return %"].iloc[0])
    else:
        latest_return = float(latest_rows["Return %"].mean())

    first_time = chart_data["Time"].min()
    last_time = chart_data["Time"].max()

    if chart_range == "LIVE":
        time_format = "%I %p"
        tick_count = 6
    elif chart_range == "1D":
        time_format = "%I %p"
        tick_count = 8
    elif chart_range == "2D":
        time_format = "%I %p"
        tick_count = 7
    elif chart_range == "1W":
        time_format = "%a"
        tick_count = 5
    else:
        time_format = "%b %d"
        tick_count = 6

    line_chart = (
        alt.Chart(chart_data)
        .mark_line(
            strokeWidth=3,
            interpolate="monotone",
        )
        .encode(
            x=alt.X(
                "Time:T",
                title=None,
                axis=alt.Axis(
                    format=time_format,
                    tickCount=tick_count,
                    grid=False,
                    labelColor="#cbd5e1",
                    labelFontSize=12,
                ),
            ),
            y=alt.Y(
                "Return %:Q",
                title=None,
                scale=alt.Scale(zero=False),
                axis=alt.Axis(
                    grid=True,
                    gridColor="#334155",
                    labelColor="#cbd5e1",
                    labelFontSize=12,
                ),
            ),
            color=alt.Color(
                "Basket:N",
                title=None,
                scale=alt.Scale(
                    domain=[
                        "Combined active basket",
                        "BUY basket",
                        "SELL short basket",
                    ],
                    range=[
                        "#047857",
                        "#1d4ed8",
                        "#b91c1c",
                    ],
                ),
                legend=alt.Legend(
                    orient="top",
                    direction="horizontal",
                    labelFontSize=13,
                    labelColor="#e5e7eb",
                    symbolSize=140,
                ),
            ),
            tooltip=[
                alt.Tooltip("Time:T", title="Time"),
                alt.Tooltip("Basket:N", title="Basket"),
                alt.Tooltip("Return %:Q", title="Return %", format=".2f"),
            ],
        )
        .properties(height=360)
    )

    zero_line = (
        alt.Chart(pd.DataFrame({"Return %": [0]}))
        .mark_rule(
            color="#64748b",
            strokeDash=[5, 5],
            strokeWidth=1.5,
        )
        .encode(y="Return %:Q")
    )

    chart = (
        line_chart + zero_line
    ).configure_view(
        strokeWidth=0
    ).configure_axis(
        domain=False,
        ticks=False,
        labelColor="#e5e7eb",
        titleColor="#e5e7eb",
    )

    return chart, latest_return, (first_time, last_time)


def style_tracker(df: pd.DataFrame):
    def color_action(value):
        value = str(value).upper()

        if value == "BUY":
            return "background-color: #047857; color: #ffffff; font-weight: 800;"
        if value == "SELL":
            return "background-color: #b91c1c; color: #ffffff; font-weight: 800;"
        if value == "WATCH":
            return "background-color: #92400e; color: #ffffff; font-weight: 800;"

        return "color: #111827;"

    def color_correct(value):
        value = str(value).upper()

        if value == "YES":
            return "background-color: #047857; color: #ffffff; font-weight: 800;"
        if value == "NO":
            return "background-color: #b91c1c; color: #ffffff; font-weight: 800;"
        if value == "N/A":
            return "background-color: #374151; color: #ffffff; font-weight: 800;"

        return "color: #111827;"

    def color_direction(value):
        value = str(value).upper()

        if value == "UP":
            return "background-color: #bbf7d0; color: #052e16; font-weight: 800;"
        if value == "DOWN":
            return "background-color: #fecaca; color: #450a0a; font-weight: 800;"
        if value == "NEUTRAL":
            return "background-color: #d1d5db; color: #111827; font-weight: 800;"

        return "color: #111827;"

    def color_edge(value):
        value = str(value).upper()

        if value == "STRONG":
            return "background-color: #1e3a8a; color: #ffffff; font-weight: 800;"
        if value == "MODERATE":
            return "background-color: #7c2d12; color: #ffffff; font-weight: 800;"
        if value == "WEAK":
            return "background-color: #374151; color: #ffffff; font-weight: 800;"

        return "color: #111827;"

    def color_regime(value):
        value = str(value).upper()

        if value == "EVENTFUL":
            return "background-color: #581c87; color: #ffffff; font-weight: 800;"
        if value == "NORMAL":
            return "background-color: #e5e7eb; color: #111827; font-weight: 800;"

        return "color: #111827;"

    def color_return(value):
        try:
            if value > 0:
                return "color: #047857; font-weight: 850;"
            if value < 0:
                return "color: #b91c1c; font-weight: 850;"
        except Exception:
            pass

        return "color: #111827;"

    styled = (
        df.style
        .map(color_action, subset=["Action"])
        .map(color_direction, subset=["Direction", "Actual Direction So Far"])
        .map(color_correct, subset=["Correct So Far"])
        .map(color_return, subset=["Change Since Monday %"])
        .format(
            {
                "Monday Reference Price": "{:.2f}",
                "Current Price": "{:.2f}",
                "Change Since Monday %": "{:.2f}%",
                "Model Score": "{:.3f}",
                "Expected Move %": "{:.2f}",
                "Setup Score": "{:.3f}",
            },
            na_rep="",
        )
    )

    if "Edge" in df.columns:
        styled = styled.map(color_edge, subset=["Edge"])

    if "Regime" in df.columns:
        styled = styled.map(color_regime, subset=["Regime"])

    return styled


def style_money(df: pd.DataFrame):
    format_map = {}

    for col in [
        "Starting Capital",
        "Current Value",
        "Dollar P/L",
        "Monday Price",
        "Current Price",
        "Monday Entry Value",
        "Current Position Value",
        "1-Share Best Correct P/L",
        "Final 1-Share P/L",
        "Wrong/Final 1-Share P/L",
        "Best Exit - Wrong P/L",
        "Final Price Used",
    ]:
        if col in df.columns:
            format_map[col] = "${:,.2f}"

    for col in [
        "Percent Return",
        "Accuracy %",
        "Avg Change Since Monday %",
        "Stock Move %",
        "Position Return %",
    ]:
        if col in df.columns:
            format_map[col] = "{:.2f}%"

    if "Shares / Shorted Shares" in df.columns:
        format_map["Shares / Shorted Shares"] = "{:,.4f}"

    def color_num(value):
        try:
            if value > 0:
                return "color: #047857; font-weight: 850;"
            if value < 0:
                return "color: #b91c1c; font-weight: 850;"
        except Exception:
            pass

        return "color: #111827; font-weight: 700;"

    def color_direction(value):
        value = str(value).upper()

        if value == "UP":
            return "background-color: #bbf7d0; color: #052e16; font-weight: 800;"
        if value == "DOWN":
            return "background-color: #fecaca; color: #450a0a; font-weight: 800;"
        if value == "NEUTRAL":
            return "background-color: #d1d5db; color: #111827; font-weight: 800;"

        return "color: #111827; font-weight: 700;"

    def color_correct(value):
        value = str(value).upper()

        if value == "YES":
            return "background-color: #047857; color: #ffffff; font-weight: 800;"
        if value == "NO":
            return "background-color: #b91c1c; color: #ffffff; font-weight: 800;"
        if value == "N/A":
            return "background-color: #374151; color: #ffffff; font-weight: 800;"

        return "color: #111827; font-weight: 700;"

    styled = df.style.format(format_map, na_rep="")

    for col in [
        "Dollar P/L",
        "Percent Return",
        "Accuracy %",
        "Avg Change Since Monday %",
        "Stock Move %",
        "Position Return %",
    ]:
        if col in df.columns:
            styled = styled.map(color_num, subset=[col])

    if "Direction" in df.columns:
        styled = styled.map(color_direction, subset=["Direction"])

    if "Correct So Far" in df.columns:
        styled = styled.map(color_correct, subset=["Correct So Far"])

    return styled



def build_excel_download(
    tracker_df: pd.DataFrame,
    weekly_path_tracker_df: pd.DataFrame,
    weekly_group_summary_df: pd.DataFrame,
    accuracy_action_df: pd.DataFrame,
    accuracy_direction_df: pd.DataFrame,
    monday_date: date,
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    weekly_truth_summary = build_weekly_truth_summary(weekly_path_tracker_df)

    dashboard_summary_df = pd.DataFrame(
        [
            ["Reference Monday", monday_date.isoformat()],
            ["Stocks checked", weekly_truth_summary["total"]],
            ["True during week", weekly_truth_summary["true_count"]],
            ["False during week", weekly_truth_summary["false_count"]],
            ["True during week %", weekly_truth_summary["true_pct"] / 100],
            ["Best exit P/L", weekly_truth_summary["best_correct_pnl"]],
            ["Wrong final P/L", weekly_truth_summary["wrong_final_pnl"]],
            ["Best exit - wrong P/L", weekly_truth_summary["perfect_exit_minus_wrong_pnl"]],
            ["Held-to-now P/L", weekly_truth_summary["final_pnl"]],
        ],
        columns=["Metric", "Value"],
    )

    export_weekly_df = weekly_path_tracker_df.copy()

    for col in ["First Correct Time", "Best Correct Time"]:
        if col in export_weekly_df.columns:
            export_weekly_df[col] = pd.to_datetime(export_weekly_df[col], errors="coerce")
            export_weekly_df[col] = export_weekly_df[col].dt.strftime("%a %I:%M %p").fillna("")

    # Full export sheet: model info + weekly truth data + current tracker fields.
    # This is the main sheet you wanted: stock name, direction, conviction, edge,
    # whether it was eventually right, when it happened, price difference since Monday, etc.
    tracker_cols = [
        "Ticker",
        "Rank",
        "Action",
        "Direction",
        "Conviction",
        "Edge",
        "Regime",
        "Model Score",
        "Expected Move %",
        "Setup Score",
        "Run Timestamp",
        "Monday Reference Price",
        "Current Price",
        "Change Since Monday %",
        "Actual Direction So Far",
        "Correct So Far",
        "Reference Price Source",
        "Price Error",
    ]

    weekly_cols = [
        "Ticker",
        "Prediction True During Week",
        "First Correct Time",
        "Best Correct Time",
        "Best Correct Price",
        "Best Correct Move %",
        "1-Share Best Correct P/L",
        "Final Price Used",
        "Final Move %",
        "Final 1-Share P/L",
    ]

    safe_tracker_cols = [c for c in tracker_cols if c in tracker_df.columns]
    safe_weekly_cols = [c for c in weekly_cols if c in export_weekly_df.columns]

    full_detail_df = tracker_df[safe_tracker_cols].merge(
        export_weekly_df[safe_weekly_cols],
        on="Ticker",
        how="left",
    )

    full_detail_df = full_detail_df.rename(
        columns={
            "Direction": "Predicted Direction",
            "Monday Reference Price": "Monday Price",
            "Change Since Monday %": "Current Move Since Monday %",
            "Correct So Far": "Correct Right Now",
            "Prediction True During Week": "Correct Eventually During Week",
            "1-Share Best Correct P/L": "Best Exit 1-Share P/L",
            "Final 1-Share P/L": "Held-To-Now 1-Share P/L",
        }
    )


    sheet_data = {
        "Full Stock Details": full_detail_df,
        "Dashboard Summary": dashboard_summary_df,
        "Weekly Truth Summary": weekly_group_summary_df,
        "Weekly Price Tracker": export_weekly_df,
        "Tracker": tracker_df,
        "Accuracy by Action": accuracy_action_df,
        "Accuracy by Direction": accuracy_direction_df,
    }

    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(color="FFFFFF", bold=True)
    normal_font = Font(color="111827")
    white_font = Font(color="FFFFFF", bold=True)

    green_fill = PatternFill("solid", fgColor="047857")
    red_fill = PatternFill("solid", fgColor="B91C1C")
    orange_fill = PatternFill("solid", fgColor="92400E")
    gray_fill = PatternFill("solid", fgColor="D1D5DB")
    dark_gray_fill = PatternFill("solid", fgColor="374151")
    blue_fill = PatternFill("solid", fgColor="1E3A8A")
    purple_fill = PatternFill("solid", fgColor="581C87")
    up_fill = PatternFill("solid", fgColor="BBF7D0")
    down_fill = PatternFill("solid", fgColor="FECACA")

    border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    money_cols = {
        "Monday Price",
        "Best Correct Price",
        "1-Share Best Correct P/L",
        "Final Price Used",
        "Final 1-Share P/L",
        "Monday Reference Price",
        "Current Price",
        "Dollar P/L",
        "Monday Entry Value",
        "Current Position Value",
        "Best Exit - Wrong P/L",
        "Wrong/Final 1-Share P/L",
        "Best Exit 1-Share P/L",
        "Held-To-Now 1-Share P/L",
    }

    percent_cols = {
        "Best Correct Move %",
        "Final Move %",
        "Change Since Monday %",
        "Expected Move %",
        "Accuracy %",
        "Avg Change Since Monday %",
        "Percent Return",
        "Current Move Since Monday %",
    }

    def safe_value(value):
        if pd.isna(value):
            return ""
        if isinstance(value, pd.Timestamp):
            return value.to_pydatetime()
        return value

    def apply_colors(cell, col_name, value):
        value_str = str(value).upper()

        if col_name in ["Prediction True During Week", "Correct So Far", "Current Correct", "Correct Eventually During Week", "Correct Right Now"]:
            if value_str == "YES":
                cell.fill = green_fill
                cell.font = white_font
            elif value_str == "NO":
                cell.fill = red_fill
                cell.font = white_font
            elif value_str == "N/A":
                cell.fill = dark_gray_fill
                cell.font = white_font

        if col_name == "Action":
            if value_str == "BUY":
                cell.fill = green_fill
                cell.font = white_font
            elif value_str == "SELL":
                cell.fill = red_fill
                cell.font = white_font
            elif value_str == "WATCH":
                cell.fill = orange_fill
                cell.font = white_font

        if col_name in ["Predicted Direction", "Direction", "Actual Direction So Far"]:
            if value_str == "UP":
                cell.fill = up_fill
                cell.font = Font(color="052E16", bold=True)
            elif value_str == "DOWN":
                cell.fill = down_fill
                cell.font = Font(color="450A0A", bold=True)
            elif value_str == "NEUTRAL":
                cell.fill = gray_fill
                cell.font = Font(color="111827", bold=True)

        if col_name == "Edge":
            if value_str == "STRONG":
                cell.fill = blue_fill
                cell.font = white_font
            elif value_str == "MODERATE":
                cell.fill = orange_fill
                cell.font = white_font
            elif value_str == "WEAK":
                cell.fill = dark_gray_fill
                cell.font = white_font

        if col_name == "Regime":
            if value_str == "EVENTFUL":
                cell.fill = purple_fill
                cell.font = white_font
            elif value_str == "NORMAL":
                cell.fill = gray_fill
                cell.font = Font(color="111827", bold=True)

        if col_name in money_cols or col_name in percent_cols:
            try:
                num = float(value)
                if num > 0:
                    cell.font = Font(color="047857", bold=True)
                elif num < 0:
                    cell.font = Font(color="B91C1C", bold=True)
            except Exception:
                pass

    for sheet_name, df_out in sheet_data.items():
        df_out = df_out.copy()
        ws = wb.create_sheet(title=sheet_name[:31])

        for col_idx, col_name in enumerate(df_out.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, (_, row) in enumerate(df_out.iterrows(), start=2):
            for col_idx, col_name in enumerate(df_out.columns, start=1):
                value = safe_value(row[col_name])
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                cell.font = normal_font
                cell.alignment = Alignment(vertical="center")
                apply_colors(cell, col_name, value)

                if col_name in money_cols and isinstance(value, (int, float)):
                    cell.number_format = '$#,##0.00;[Red]-$#,##0.00'
                elif col_name in percent_cols and isinstance(value, (int, float)):
                    cell.number_format = '0.00"%"'

        if sheet_name == "Dashboard Summary":
            for row_idx in range(2, ws.max_row + 1):
                metric = ws.cell(row=row_idx, column=1).value
                if metric == "True during week %":
                    ws.cell(row=row_idx, column=2).number_format = "0.00%"
                if metric in [
                    "Best exit P/L",
                    "Wrong final P/L",
                    "Best exit - wrong P/L",
                    "Held-to-now P/L",
                ]:
                    ws.cell(row=row_idx, column=2).number_format = '$#,##0.00;[Red]-$#,##0.00'

        ws.freeze_panes = "A2"
        if ws.max_column > 0 and ws.max_row > 1:
            ws.auto_filter.ref = ws.dimensions

        for col_idx, col_name in enumerate(df_out.columns, start=1):
            values = [str(col_name)]
            for row_idx in range(2, min(ws.max_row, 200) + 1):
                values.append(str(ws.cell(row=row_idx, column=col_idx).value or ""))
            width = min(max(max(len(v) for v in values) + 2, 12), 36)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


st.title("Manual Weekly Stock Signal Tracker")
st.caption("Tracks this week's model output against this week's Monday reference price, not daily percent change.")

if "raw_model_output" not in st.session_state:
    st.session_state["raw_model_output"] = DEFAULT_TEXT

tab_dashboard, tab_input = st.tabs(["Dashboard", "Paste model output"])

with tab_input:
    st.session_state["raw_model_output"] = st.text_area(
        "Paste model output",
        value=st.session_state["raw_model_output"],
        height=360,
    )
    st.info("After editing this box, go back to the Dashboard tab. The tracker updates automatically.")

model_df = parse_model_output(st.session_state["raw_model_output"])
monday_date = this_weeks_monday()

with tab_dashboard:
    st.caption(f"Reference period: this week's Monday ({monday_date}) to now.")

    # Position tables use a simple one-share-per-stock model.

    if model_df.empty:
        st.warning("No valid rows found. Paste your model output in the second tab.")
        st.stop()

    tickers = tuple(model_df["Ticker"].dropna().astype(str).str.upper().unique().tolist())
    model_rows = tuple(
        model_df[["Ticker", "Action", "Direction"]]
        .dropna()
        .astype(str)
        .itertuples(index=False, name=None)
    )

    with st.spinner("Fetching prices and calculating performance..."):
        price_df = fetch_prices(tickers, monday_date)
        tracker_df = add_tracking_columns(model_df, price_df)
        portfolio_chart_df = fetch_position_return_series(model_rows, monday_date)
        weekly_path_price_df = fetch_week_price_paths(tickers, monday_date)
        weekly_path_tracker_df = build_weekly_price_tracker(tracker_df, weekly_path_price_df)

    active_valid = tracker_df[
        (tracker_df["Action"].isin(["BUY", "SELL"]))
        & (tracker_df["Correct So Far"].isin(["YES", "NO"]))
    ].copy()

    active_correct = int((active_valid["Correct So Far"] == "YES").sum()) if not active_valid.empty else 0
    active_total = int(len(active_valid))
    active_accuracy = active_correct / active_total * 100 if active_total else 0

    what_if_df = build_what_if(tracker_df)
    combined_row = what_if_df[what_if_df["Strategy"] == "Combined active calls: 1 share each BUY/UP + SELL/DOWN"]

    if not combined_row.empty and pd.notna(combined_row["Dollar P/L"].iloc[0]):
        combined_pnl = float(combined_row["Dollar P/L"].iloc[0])
        combined_return = float(combined_row["Percent Return"].iloc[0])
    else:
        combined_pnl = 0.0
        combined_return = 0.0

    weekly_truth_summary = build_weekly_truth_summary(weekly_path_tracker_df)

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Active calls", active_total)
    c2.metric("Active correct now", active_correct)
    c3.metric(
        "True during week",
        f"{weekly_truth_summary['true_count']} / {weekly_truth_summary['total']}",
    )
    c4.metric("True during week %", f"{weekly_truth_summary['true_pct']:.1f}%")

    c5, c6, c7, c8 = st.columns(4)
    c5.metric(
        "Best exit P/L",
        f"${weekly_truth_summary['best_correct_pnl']:,.2f}",
        "Correct calls only",
    )
    c6.metric(
        "Wrong final P/L",
        f"${weekly_truth_summary['wrong_final_pnl']:,.2f}",
        "Never-correct calls",
    )
    c7.metric(
        "Best exit - wrong",
        f"${weekly_truth_summary['perfect_exit_minus_wrong_pnl']:,.2f}",
        "Sell/cover at best if right",
    )
    c8.metric(
        "Held-to-now P/L",
        f"${weekly_truth_summary['final_pnl']:,.2f}",
        "Latest price result",
    )

    st.subheader("Portfolio performance")

    chart_range = st.radio(
        "Range",
        ["LIVE", "1D", "2D", "1W"],
        horizontal=True,
        label_visibility="collapsed",
    )

    filtered_chart_df = filter_chart_range(portfolio_chart_df, chart_range)

    st.caption(
        "Equal-weight basket returns. BUY/UP is treated as long. SELL/DOWN is treated as short. WATCH is ignored."
    )

    if filtered_chart_df.empty:
        st.info("No portfolio chart data available yet.")
    else:
        chart, latest_return, time_window = make_portfolio_chart(filtered_chart_df, chart_range)

        if chart is None:
            st.info("No portfolio chart data available yet.")
        else:
            first_time, last_time = time_window
            change_class = "chart-change-green" if latest_return >= 0 else "chart-change-red"
            change_sign = "+" if latest_return >= 0 else ""

            st.markdown('<div class="chart-wrap">', unsafe_allow_html=True)

            st.markdown(
                f"""
                <div class="chart-title-row">
                    <div>
                        <div class="chart-big-number">{change_sign}{latest_return:.2f}%</div>
                        <div class="{change_class}">Combined active basket</div>
                    </div>
                    <div class="chart-small-note">
                        {first_time.strftime("%b %d, %I:%M %p")} → {last_time.strftime("%b %d, %I:%M %p")}
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )

            st.altair_chart(chart, use_container_width=True)

            st.markdown("</div>", unsafe_allow_html=True)

    st.subheader("Weekly price-path tracker")
    st.caption(
        "This checks whether each prediction became true at any point during the week. "
        "It removes current price from this table and focuses on when the call was correct, the best correct price, and the 1-share profit/loss."
    )

    if weekly_path_tracker_df.empty:
        st.info("No weekly price-path data available yet.")
    else:
        true_week_df = weekly_path_tracker_df[
            weekly_path_tracker_df["Prediction True During Week"] == "YES"
        ].copy()

        false_week_df = weekly_path_tracker_df[
            weekly_path_tracker_df["Prediction True During Week"] == "NO"
        ].copy()

        all_week_df = weekly_path_tracker_df.copy()

        st.subheader("Weekly truth summary")

        weekly_group_summary_df = build_weekly_truth_group_summary(weekly_path_tracker_df)

        st.dataframe(
            style_money(weekly_group_summary_df),
            use_container_width=True,
            hide_index=True,
        )

        true_tab, false_tab, all_tab = st.tabs(
            ["True during week", "False during week", "All stocks"]
        )

        display_cols = [
            "Ticker",
            "Action",
            "Predicted Direction",
            "Prediction True During Week",
            "Monday Price",
            "First Correct Time",
            "Best Correct Time",
            "Best Correct Price",
            "Best Correct Move %",
            "1-Share Best Correct P/L",
            "Final Price Used",
            "Final Move %",
            "Final 1-Share P/L",
        ]

        with true_tab:
            st.caption(
                "These calls were correct at least once during the week. "
                "Best Correct Price is the best price in the predicted direction."
            )
            st.dataframe(
                style_weekly_path_tracker(true_week_df[display_cols]),
                use_container_width=True,
                hide_index=True,
            )

        with false_tab:
            st.caption(
                "These calls never became correct during the week. "
                "Best-correct P/L stays at $0 because the predicted move never happened."
            )
            st.dataframe(
                style_weekly_path_tracker(false_week_df[display_cols]),
                use_container_width=True,
                hide_index=True,
            )

        with all_tab:
            st.caption(
                "All stocks together, with final price kept only to show what the total ending result would be."
            )
            st.dataframe(
                style_weekly_path_tracker(all_week_df[display_cols]),
                use_container_width=True,
                hide_index=True,
            )

        st.subheader("Position details")
    st.caption(
        "Assumes 1 share bought for each BUY/UP call and 1 share shorted for each SELL/DOWN call at Monday's reference price."
    )

    detail_buy, detail_sell, detail_combined = st.tabs(
        ["BUY/UP positions", "SELL/DOWN short positions", "Combined active positions"]
    )

    with detail_buy:
        st.dataframe(
            style_money(build_what_if_positions(tracker_df, "BUY")),
            use_container_width=True,
            hide_index=True,
        )

    with detail_sell:
        st.dataframe(
            style_money(build_what_if_positions(tracker_df, "SELL")),
            use_container_width=True,
            hide_index=True,
        )

    with detail_combined:
        st.dataframe(
            style_money(build_what_if_positions(tracker_df, "COMBINED")),
            use_container_width=True,
            hide_index=True,
        )

    st.subheader("Robinhood bearish trade calculators")
    st.caption(
        "For SELL/DOWN calls. The paper short calculator shows classic short math. "
        "The put calculator is the cleaner Robinhood-style bearish setup if your account is approved for options."
    )

    short_df = build_short_calculator_table(tracker_df)

    if short_df.empty:
        st.info("No valid SELL/DOWN stocks available yet.")
    else:
        calc_tab_short, calc_tab_put = st.tabs(["Paper short calculator", "Put option calculator"])

        with calc_tab_short:
            selected_short_ticker = st.selectbox(
                "Select stock to short",
                short_df["Ticker"].tolist(),
                key="short_calc_ticker",
            )

            selected_row = short_df[short_df["Ticker"] == selected_short_ticker].iloc[0]

            monday_price = float(selected_row["Monday Reference Price"])
            current_price = float(selected_row["Current Price"])

            col_a, col_b, col_c = st.columns(3)

            with col_a:
                short_amount = st.number_input(
                    "Dollar amount to short",
                    min_value=0.0,
                    value=1000.0,
                    step=100.0,
                    key="short_calc_amount",
                )

            with col_b:
                target_cover_price = st.number_input(
                    "Target cover price",
                    min_value=0.01,
                    value=current_price,
                    step=1.0,
                    key="short_calc_target_price",
                )

            with col_c:
                st.metric("Monday short price", f"${monday_price:,.2f}")
                st.metric("Current price", f"${current_price:,.2f}")

            shares_shorted = short_amount / monday_price if monday_price else 0.0

            current_pnl = shares_shorted * (monday_price - current_price)
            target_pnl = shares_shorted * (monday_price - target_cover_price)

            current_return_pct = (current_pnl / short_amount * 100) if short_amount else 0.0
            target_return_pct = (target_pnl / short_amount * 100) if short_amount else 0.0

            calc_cols = st.columns(4)

            calc_cols[0].metric("Shares shorted", f"{shares_shorted:,.4f}")
            calc_cols[1].metric("P/L if covered now", f"${current_pnl:,.2f}", f"{current_return_pct:.2f}%")
            calc_cols[2].metric("P/L at target cover", f"${target_pnl:,.2f}", f"{target_return_pct:.2f}%")
            calc_cols[3].metric("Target cover price", f"${target_cover_price:,.2f}")

            st.caption(
                "Paper short logic: short at the Monday price, then buy back later. "
                "If the stock falls, the short gains. If the stock rises, the short loses."
            )

        with calc_tab_put:
            selected_put_ticker = st.selectbox(
                "Select stock for put option",
                short_df["Ticker"].tolist(),
                key="put_calc_ticker",
            )

            selected_put_row = short_df[short_df["Ticker"] == selected_put_ticker].iloc[0]

            monday_price = float(selected_put_row["Monday Reference Price"])
            current_price = float(selected_put_row["Current Price"])

            put_col_a, put_col_b, put_col_c, put_col_d = st.columns(4)

            with put_col_a:
                contracts = st.number_input(
                    "Contracts",
                    min_value=1,
                    value=1,
                    step=1,
                    key="put_contracts",
                )

            with put_col_b:
                premium_paid = st.number_input(
                    "Premium paid per share",
                    min_value=0.01,
                    value=1.00,
                    step=0.05,
                    key="put_premium_paid",
                )

            with put_col_c:
                current_premium = st.number_input(
                    "Current premium per share",
                    min_value=0.00,
                    value=1.00,
                    step=0.05,
                    key="put_current_premium",
                )

            with put_col_d:
                target_premium = st.number_input(
                    "Target premium per share",
                    min_value=0.00,
                    value=1.50,
                    step=0.05,
                    key="put_target_premium",
                )

            total_cost = contracts * premium_paid * 100
            current_value = contracts * current_premium * 100
            target_value = contracts * target_premium * 100

            current_pnl = current_value - total_cost
            target_pnl = target_value - total_cost

            current_return_pct = (current_pnl / total_cost * 100) if total_cost else 0.0
            target_return_pct = (target_pnl / total_cost * 100) if total_cost else 0.0

            put_metrics = st.columns(5)

            put_metrics[0].metric("Stock Monday price", f"${monday_price:,.2f}")
            put_metrics[1].metric("Stock current price", f"${current_price:,.2f}")
            put_metrics[2].metric("Total option cost", f"${total_cost:,.2f}")
            put_metrics[3].metric("P/L now", f"${current_pnl:,.2f}", f"{current_return_pct:.2f}%")
            put_metrics[4].metric("P/L at target", f"${target_pnl:,.2f}", f"{target_return_pct:.2f}%")

            st.caption(
                "Put option logic: one contract usually controls 100 shares. "
                "If the put premium rises after the stock falls, the option position gains. "
                "Max loss for a bought put is the premium paid, but options can expire worthless."
            )

    st.subheader("Tracker")
    st.dataframe(style_tracker(tracker_df), use_container_width=True, hide_index=True)

    col1, col2 = st.columns(2)
    with col1:
        st.subheader("Accuracy by action")
        st.dataframe(style_money(build_summary(tracker_df, "Action")), use_container_width=True, hide_index=True)

    with col2:
        st.subheader("Accuracy by direction")
        st.dataframe(style_money(build_summary(tracker_df, "Direction")), use_container_width=True, hide_index=True)

    st.subheader("Download results")

    weekly_group_summary_df = build_weekly_truth_group_summary(weekly_path_tracker_df)
    accuracy_action_df = build_summary(tracker_df, "Action")
    accuracy_direction_df = build_summary(tracker_df, "Direction")

    excel_bytes = build_excel_download(
        tracker_df=tracker_df,
        weekly_path_tracker_df=weekly_path_tracker_df,
        weekly_group_summary_df=weekly_group_summary_df,
        accuracy_action_df=accuracy_action_df,
        accuracy_direction_df=accuracy_direction_df,
        monday_date=monday_date,
    )

    st.download_button(
        "Download full color-coded Excel tracker",
        data=excel_bytes,
        file_name=f"manual_weekly_tracker_{monday_date}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
